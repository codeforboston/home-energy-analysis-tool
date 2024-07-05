"""
This module is designed to be called at "build time" to read example excel files and generate the following test input files:
    A summary.json file containing info from the Summary tab of the spreadsheet
    An energy usage csv file containing info from either the Natural Gas or Oil-Propane tab of the spreadsheet
We intend to call this code infrequently, rather than with each test run because the excel sheets are not changed often and
the effort to ingest the excel sheets inline significantly slows down our tests by 10x

Test example spreadsheets are provided in separate household folders within the "cases/examples" directory
Each folder contains an excel file (named "Heat Load Analysis Beta 7.xlsx") which specifies the example inputs
Once this module is run, each folder will contain three files, the original excel, a summary.json, and a fuel-specific csv file

NOTE: Due to the variety of types in the Summary and Fuel specific worksheets, the workbook object is purposely of type Any.
"""

import csv
import json
import os
import pathlib
import re
from pathlib import Path
from typing import Any

import openpyxl

ROOT_DIR = pathlib.Path(__file__).parent / "cases" / "examples"

# Identify all available example data directories
INPUT_DATA_DIRS = list(os.walk(ROOT_DIR))


def generate_summary_json(workbook: Any, working_directory: Path) -> str:
    """
    Read the heat load analysis spreadsheet and write information from the "Summary" tab into a json file
    We do this so our test runners use the json file for faster processing of our example data
    Return the fuel_type to the caller so that our csv generator function can read the fuel billing information
    """
    # Choose the Summary worksheet as the source for summary.json data
    worksheet = workbook["Summary"]
    data = {}

    for row in worksheet.iter_rows(min_row=1, max_col=2, values_only=True):
        match row[0], row[1]:
            case ["Local Weather Station", value]:
                data["local_weather_station"] = value
            case ["Design temperature override", value]:
                data["design_temperature_override"] = value or None
            case ["Living area", value]:
                data["living_area"] = value
            case ["Fuel type", value]:
                if "Gas" in value:
                    data["fuel_type"] = fuel_type = "GAS"
                elif "Oil" in value:
                    data["fuel_type"] = fuel_type = "OIL"
                else:
                    raise ValueError("Expected: 'Gas' or 'Oil'; Instead found:", value)
            case ["Heating system efficiency", value]:
                data["heating_system_efficiency"] = round(value, 2)
            case ["Other fuel usage (DHW, cooking)", value]:
                data["other_fuel_usage"] = round(value or 0, 2)
            case ["Other fuel usage OVERRIDE", value]:
                data["other_fuel_usage_override"] = value or None
            case ["Thermostat set point", value]:
                data["thermostat_set_point"] = value
            case ["Setback temp (nights,…)", value]:
                data["setback_temperature"] = value or None
            case ["Setback hours per day", value]:
                data["setback_hours_per_day"] = value or None
            case ["Estimated balance point, Tbp", value]:
                data["estimated_balance_point"] = value
            case ["Balance point sensitivity", value]:
                data["balance_point_sensitivity"] = value
            case ["Average Indoor Temperature, Ti", value]:
                data["average_indoor_temperature"] = round(value, 1)
            case ["Difference between Ti and Tbp", value]:
                data["difference_between_ti_and_tbp"] = round(value, 1)
            case ["Design temperature", value]:
                data["design_temperature"] = round(value, 1)
            case ["Whole-home UA", value]:
                data["whole_home_heat_loss_rate"] = round(value, 0)
            case ["Standard deviation of UA", value]:
                data["standard_deviation_of_heat_loss_rate"] = round(value, 4)
            case ["Avg Heat Load @ Design temp.", value]:
                data["average_heat_load"] = round(value, 0)
            case ["Max Heat Load @ Design temp.", value]:
                data["maximum_heat_load"] = round(value, 0)
                break

    # Now that we have accumulated all the relevant fields into a data dictionary, write out summary.json
    with open(working_directory / "summary.json", "w") as json_file:
        json.dump(data, json_file, indent=4)

    # Return the fuel type we found in the Summary worksheet so we can operate on the correct worksheet next
    return fuel_type


def generate_billing_record_input_csv(
    workbook: Any, fuel_type: str, working_directory: Path
) -> None:
    """
    Read the heat load analysis spreadsheet and write data from the appropriate "fuel type" worksheet into a csv file
    We do this so our test runners use the csv file for faster processing of our example data
    """
    # Choose the appropriate fuel-type worksheet, set header and data row locations and output filename for each type
    if fuel_type == "GAS":
        output_file_path = working_directory / "natural-gas.csv"
        worksheet = workbook["Natural Gas"]
        header_row = 4
        billing_row = 5
        # Setup the column header names in the appropriate positions for the Natural Gas worksheet
        header = [
            "start_date",
            "end_date",
            "days_in_bill",
            "usage",
            "inclusion_override",
            "inclusion_code",
            "avg_daily_usage",
            "daily_htg_usage",
        ]
    elif fuel_type == "OIL":
        output_file_path = working_directory / "oil-propane.csv"
        worksheet = workbook["Oil-Propane"]
        header_row = 5
        billing_row = 7
        # Setup the column header names in the appropriate positions for the Oil-Propane worksheet
        header = [
            "start_date",
            "end_date",
            "usage",
            "days_in_bill",
            "inclusion_code",
            "inclusion_override",
            "avg_daily_usage",
            "daily_htg_usage",
        ]
    else:
        raise ValueError("Expected: 'GAS' or 'OIL'; Instead found:", fuel_type)

    # Read just the column headers for processing into more readable forms for our csv output
    row = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=1,
            max_col=16,
            values_only=False,
        )
    )

    # Find just the digits portion of the hdd, ua, and ua_sensitivity to generate unique headers
    for i in range(3):
        temp = re.search(r"[-+]?\d*\.\d+|\d+", row[8 + i].value)
        if temp is None:
            raise ValueError("Expected to find a temperature, found none")
        header.append("hdd_at_" + str(temp.group()) + "f")

    for i in range(3):
        temp = re.search(r"[-+]?\d*\.\d+|\d+", row[11 + i].value)
        if temp is None:
            raise ValueError("Expected to find a temperature, found none")
        header.append("ua_at_" + str(temp.group()) + "f")

    for i in range(2):
        temp = re.search(r"[-+]?\d*\.\d+|\d+", row[14 + i].value)
        if temp is None:
            raise ValueError("Expected to find a temperature, found none")
        header.append("ua_sensitivity_at_" + str(temp.group()) + "_therms")

    # Write the header information first, then copy billing data row by row into the csv file
    with open(output_file_path, mode="w", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(header)

        # Now extract and write all the row data from excel to csv
        for row in worksheet.iter_rows(
            min_row=billing_row, max_col=16, values_only=True
        ):
            if row[0] == None:
                break
            writer.writerow(row)


if __name__ == "__main__":
    # For each example folder, read the excel sheet and write summary.json and "fuel_type".csv
    for root, dirs, files in os.walk(ROOT_DIR):
        working_directory = Path(root)
        print("found directory =", working_directory)
        try:
            workbook = openpyxl.load_workbook(
                filename=working_directory / "Heat Load Analysis Beta 7.xlsx",
                data_only=True,
            )
        except FileNotFoundError as fnf_error:
            # It's ok to iterate over directories that do not contain an excel file
            continue
        print("processing directory =", working_directory)
        fuel_type = generate_summary_json(workbook, working_directory)
        generate_billing_record_input_csv(workbook, fuel_type, working_directory)
        workbook.close()
        del workbook
